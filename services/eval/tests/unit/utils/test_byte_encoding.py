import io
import json
import pytest
import pandas as pd
import yaml
from openpyxl import Workbook
from llm_eval.utils.byte_encoding import to_bytes_with_format


def test_to_bytes_with_format_json() -> None:
    data = {"key": "value"}
    expected_bytes = json.dumps(data, indent=2).encode("utf-8")
    expected_type = "application/json"

    result = to_bytes_with_format(data, "json")

    assert result == (expected_bytes, expected_type)


def test_to_bytes_with_format_csv() -> None:
    data = [{"col1": "value1", "col2": "value2"}]
    df = pd.DataFrame(data)
    csv_buffer = io.StringIO()
    df.to_csv(csv_buffer, index=False)
    expected_bytes = csv_buffer.getvalue().encode("utf-8")
    expected_type = "text/csv"

    result = to_bytes_with_format(data, "csv")

    assert result == (expected_bytes, expected_type)


def test_to_bytes_with_format_yaml() -> None:
    data = {"key": "value"}
    expected_bytes = yaml.dump(data, default_flow_style=False, sort_keys=False).encode(
        "utf-8"
    )
    expected_type = "application/x-yaml"

    result = to_bytes_with_format(data, "yaml")

    assert result == (expected_bytes, expected_type)


def test_to_bytes_with_format_xlsx() -> None:
    wb = Workbook()
    ws = wb.active

    assert ws is not None

    ws["A1"] = "key"
    ws["B1"] = "value"
    ws["A2"] = "example"
    ws["B2"] = "data"

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    expected_bytes = excel_buffer.getvalue()
    expected_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    result = to_bytes_with_format(wb, "xlsx")

    assert result == (expected_bytes, expected_type)


def test_to_bytes_with_format_invalid_format() -> None:
    data = {"key": "value"}

    with pytest.raises(ValueError, match="Unsupported format: invalid"):
        to_bytes_with_format(data, "invalid")  # type: ignore


def test_to_bytes_with_format_invalid_xlsx_data() -> None:
    data = {"key": "value"}

    result = to_bytes_with_format(data, "xlsx")

    assert result is None
