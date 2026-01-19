import json
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl as excel
import pytest
from sqlalchemy.ext.asyncio import AsyncSession

from llm_eval.qa_catalog.logic.download import (
    _qa_pairs_to_excel,
    find_all_qa_pairs,
    get_single_catalog_bytes,
    handle_catalog_download,
    qa_pairs_to_bytes_with_format,
)
from llm_eval.qa_catalog.models import (
    DownloadQACatalogOptions,
    QAPair,
)


@pytest.mark.asyncio
@patch(
    "llm_eval.qa_catalog.logic.download._qa_pairs_to_excel",
    return_value=b"excel_data",
)
@patch(
    "llm_eval.qa_catalog.logic.download.to_bytes_with_format",
    return_value=(
        b"bytes_data",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ),
)
async def test_qa_pairs_to_bytes_with_format(
    mock_to_bytes_with_format: MagicMock, mock_qa_pairs_to_excel: MagicMock
) -> None:
    qa_pairs = [
        QAPair(
            id="pair1",
            question="What is AI?",
            expected_output="Artificial Intelligence",
            contexts=["context1"],
            meta_data={"source": "source1"},
        ),
        QAPair(
            id="pair2",
            question="What is ML?",
            expected_output="Machine Learning",
            contexts=["context2"],
            meta_data={"source": "source2"},
        ),
    ]
    result = qa_pairs_to_bytes_with_format(qa_pairs, "xlsx")
    assert result is not None
    assert result[0] == b"bytes_data"
    assert (
        result[1] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@pytest.mark.asyncio
@patch(
    "llm_eval.qa_catalog.logic.download.find_all_qa_pairs",
    return_value=[
        QAPair(
            id="pair1",
            question="What is AI?",
            expected_output="Artificial Intelligence",
            contexts=["context1"],
            meta_data={"source": "source1"},
        ),
        QAPair(
            id="pair2",
            question="What is ML?",
            expected_output="Machine Learning",
            contexts=["context2"],
            meta_data={"source": "source2"},
        ),
    ],
)
@patch(
    "llm_eval.qa_catalog.logic.download.qa_pairs_to_bytes_with_format",
    return_value=(
        b"bytes_data",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ),
)
async def test_get_single_catalog_bytes(
    mock_qa_pairs_to_bytes_with_format: MagicMock, mock_find_all_qa_pairs: MagicMock
) -> None:
    db = AsyncMock(AsyncSession)
    catalog_id = "test_catalog_id"
    format = "xlsx"

    result = await get_single_catalog_bytes(db, catalog_id, format)
    assert result is not None
    assert result[0] == b"bytes_data"
    assert (
        result[1] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@pytest.mark.asyncio
@patch(
    "llm_eval.qa_catalog.logic.download.get_single_catalog_bytes",
    return_value=(
        b"bytes_data",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ),
)
@patch("llm_eval.qa_catalog.logic.download.datetime")
async def test_handle_catalog_download(
    mock_datetime: MagicMock, mock_get_single_catalog_bytes: MagicMock
) -> None:
    db = AsyncMock(AsyncSession)
    options = DownloadQACatalogOptions(
        parent_catalog_id="test_catalog_id", format="xlsx"
    )

    result = await handle_catalog_download(db, options)
    assert result is not None
    assert result.download_url.startswith(
        "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
    )


@pytest.mark.asyncio
@patch(
    "llm_eval.qa_catalog.logic.download.find_qa_catalog",
    return_value={"id": "test_catalog_id"},
)
@patch(
    "llm_eval.qa_catalog.logic.download._all_pairs",
    return_value=[
        QAPair(
            id="pair1",
            question="What is AI?",
            expected_output="Artificial Intelligence",
            contexts=["context1"],
            meta_data={"source": "source1"},
        ),
        QAPair(
            id="pair2",
            question="What is ML?",
            expected_output="Machine Learning",
            contexts=["context2"],
            meta_data={"source": "source2"},
        ),
    ],
)
async def test_find_all_qa_pairs(
    mock_all_pairs: MagicMock, mock_find_qa_catalog: MagicMock
) -> None:
    db = AsyncMock(AsyncSession)
    catalog_id = "test_catalog_id"

    result = await find_all_qa_pairs(db, catalog_id)
    assert result is not None
    assert len(result) == 2
    assert result[0].id == "pair1"
    assert result[1].id == "pair2"


@pytest.mark.asyncio
@patch("llm_eval.qa_catalog.logic.download.find_qa_catalog", return_value=None)
async def test_find_all_qa_pairs_no_catalog(mock_find_qa_catalog: MagicMock) -> None:
    db = AsyncMock(AsyncSession)
    catalog_id = "non_existent_catalog_id"

    result = await find_all_qa_pairs(db, catalog_id)
    assert result is None


def test_qa_pairs_to_bytes_with_format_invalid_format() -> None:
    qa_pairs = [
        QAPair(
            id="pair1",
            question="What is AI?",
            expected_output="Artificial Intelligence",
            contexts=["context1"],
            meta_data={"source": "source1"},
        ),
        QAPair(
            id="pair2",
            question="What is ML?",
            expected_output="Machine Learning",
            contexts=["context2"],
            meta_data={"source": "source2"},
        ),
    ]
    with pytest.raises(ValueError, match="Unsupported format: invalid_format"):
        qa_pairs_to_bytes_with_format(qa_pairs, "invalid_format")  # type: ignore


@pytest.mark.asyncio
@patch("llm_eval.qa_catalog.logic.download.find_qa_catalog", return_value=None)
async def test_handle_catalog_download_no_parent_catalog(
    mock_find_qa_catalog: MagicMock,
) -> None:
    db = AsyncMock(AsyncSession)
    options = DownloadQACatalogOptions(
        parent_catalog_id="non_existent_catalog_id", format="xlsx", include_all=True
    )

    result = await handle_catalog_download(db, options)
    assert result is None


@pytest.mark.asyncio
@patch(
    "llm_eval.qa_catalog.logic.download.get_single_catalog_bytes",
    return_value=(
        b"bytes_data",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ),
)
async def test_handle_catalog_download_multiple_catalogs(
    mock_get_single_catalog_bytes: MagicMock,
) -> None:
    db = AsyncMock(AsyncSession)
    options = DownloadQACatalogOptions(
        parent_catalog_id="test_catalog_id",
        format="xlsx",
        version_ids=["version1", "version2"],
    )

    result = await handle_catalog_download(db, options)
    assert result is not None
    assert result.download_url.startswith("data:application/zip;base64,")
    assert result.filename.endswith(".zip")


def test__qa_pairs_to_excel_contains_headers_and_data() -> None:
    # Create sample QAPair objects
    pair1 = QAPair(
        id="pair1",
        question="What is AI?",
        expected_output="Artificial Intelligence",
        contexts=["context1", "context1.1"],
        meta_data={"source": "source1"},
    )
    pair2 = QAPair(
        id="pair2",
        question="What is ML?",
        expected_output="Machine Learning",
        contexts=["context2"],
        meta_data={"source": "source2"},
    )
    qa_pairs = [pair1, pair2]

    # Generate workbook using the private function
    wb = _qa_pairs_to_excel(qa_pairs)
    assert wb is not None, "The workbook should be created"
    ws = wb.active

    assert ws is not None

    # Get header row from worksheet
    header_row = [cell.value for cell in ws[1]]
    # Expected headers are from model_dump of the first QAPair
    expected_headers = list(pair1.model_dump().keys())
    assert (
        header_row == expected_headers
    ), "Headers in excel should match keys of QAPair model"

    # Prepare expected row data for pair1 with transformed fields
    p1_data = pair1.model_dump()
    p1_data["contexts"] = "\n".join(p1_data["contexts"])
    p1_data["meta_data"] = json.dumps(p1_data["meta_data"])
    expected_row_1 = list(p1_data.values())

    # Extract row 2 values from sheet
    row_2 = [cell.value for cell in ws[2]]
    assert row_2 == expected_row_1, "Row for first QA pair does not match expected data"

    # Prepare expected row data for pair2 with transformed fields
    p2_data = pair2.model_dump()
    p2_data["contexts"] = "\n".join(p2_data["contexts"])
    p2_data["meta_data"] = json.dumps(p2_data["meta_data"])
    expected_row_2 = list(p2_data.values())

    # Extract row 3 values from sheet
    row_3 = [cell.value for cell in ws[3]]
    assert (
        row_3 == expected_row_2
    ), "Row for second QA pair does not match expected data"


def test__qa_pairs_to_excel_returns_workbook_instance() -> None:
    pair = QAPair(
        id="test_pair",
        question="Sample?",
        expected_output="Yes",
        contexts=["sample"],
        meta_data={"key": "value"},
    )
    qa_pairs = [pair]
    wb = _qa_pairs_to_excel(qa_pairs)
    assert isinstance(
        wb, excel.Workbook
    ), "Returned object must be an openpyxl Workbook"
