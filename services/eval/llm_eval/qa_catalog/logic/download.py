import base64
from datetime import datetime
import io
from typing import List
import zipfile
import json

import openpyxl as excel
from loguru import logger
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from llm_eval.qa_catalog.db.find_qa_catalog import find_qa_catalog
from llm_eval.qa_catalog.db.find_qa_pairs import find_qa_pairs
from llm_eval.qa_catalog.logic.revision_history import (
    create_qa_catalog_revision_history,
)
from llm_eval.qa_catalog.models import (
    DownloadQACatalogOptions,
    DownloadQACatalogResponse,
    QAPair,
    SupportedQACatalogDownloadFormat,
)
from llm_eval.utils.byte_encoding import (
    to_bytes_with_format,
)


async def _all_pairs(
    db: AsyncSession,
    catalog_id: str,
    limit: int = 100,
    offset: int = 0,
) -> list[QAPair]:
    all_pairs: List[QAPair] = []
    while True:
        paged_pairs = await find_qa_pairs(db, catalog_id, limit, offset)
        all_pairs.extend(QAPair.model_validate(p) for p in paged_pairs)

        if len(paged_pairs) < limit:  # last page reached
            break

        offset += limit

    return [QAPair.model_validate(p) for p in all_pairs]


async def find_all_qa_pairs(db: AsyncSession, catalog_id: str) -> list[QAPair] | None:
    catalog = await find_qa_catalog(db, catalog_id)
    if not catalog:
        return None
    qa_pairs = await _all_pairs(db, catalog_id)

    return qa_pairs


def _qa_pairs_to_excel(
    qa_pairs: list[QAPair],
) -> excel.Workbook | None:
    wb = excel.Workbook()
    ws = wb.active
    if not ws:
        return None

    dict_pairs = [p.model_dump() for p in qa_pairs]

    formatted_dict_pairs = [
        {
            **p,
            "contexts": "\n".join(p["contexts"]),
            "meta_data": json.dumps(p["meta_data"]),
        }
        for p in dict_pairs
    ]
    # Write headers
    headers = list(formatted_dict_pairs[0].keys())
    ws.append(headers)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Write data
    for r in formatted_dict_pairs:
        if _r := list(r.values()):
            ws.append(_r)
            # Center-align all cells in the row
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0

        if column[0].column is None:
            continue

        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                logger.error(e)
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    return wb


def qa_pairs_to_bytes_with_format(
    qa_pairs: list[QAPair], format: SupportedQACatalogDownloadFormat
) -> tuple[bytes, str] | None:
    if format == "xlsx":
        data = _qa_pairs_to_excel(qa_pairs)
    else:
        data = [p.model_dump() for p in qa_pairs]

    return to_bytes_with_format(data, format)


async def get_single_catalog_bytes(
    db: AsyncSession, catalog_id: str, format: SupportedQACatalogDownloadFormat
) -> tuple[bytes, str] | None:
    qa_pairs = await find_all_qa_pairs(db, catalog_id)
    if not qa_pairs:
        return None

    data_tuple = qa_pairs_to_bytes_with_format(qa_pairs, format)

    return data_tuple


async def handle_catalog_download(
    db: AsyncSession, options: DownloadQACatalogOptions
) -> DownloadQACatalogResponse | None:
    ids_to_download: list[str] = []
    if options.include_all:
        parent_catalog = await find_qa_catalog(db, options.parent_catalog_id)
        if not parent_catalog:
            return None
        history = await create_qa_catalog_revision_history(db, parent_catalog)
        ids_to_download = [v.version_id for v in history.versions]
    elif options.version_ids:
        ids_to_download = list(
            set(options.version_ids).union({options.parent_catalog_id})
        )
    else:
        ids_to_download = [options.parent_catalog_id]

    all_bytes: list[tuple[str, bytes]] = []
    data_type = None
    for catalog_id in ids_to_download:
        result = await get_single_catalog_bytes(db, catalog_id, options.format)
        if result:
            all_bytes.append((catalog_id, result[0]))
            data_type = result[1]

    if len(all_bytes) > 1:
        b = io.BytesIO()
        z = zipfile.ZipFile(b, "w")

        for catalog_id, catalog_bytes in all_bytes:
            info = zipfile.ZipInfo(f"{catalog_id}.{options.format}")
            z.writestr(info, catalog_bytes)

        z.close()
        b.flush()
        b.seek(0)
        data_bytes = b.getvalue()
        data_type = "application/zip"
        archive_date = datetime.now().strftime("catalog_archive-%d/%m/%Y-%H:%M:%S")
        filename = f"{archive_date}.zip"
    else:
        data_bytes = all_bytes[0][1]
        filename = f"{all_bytes[0][0]}.{options.format}"

    data = base64.b64encode(data_bytes).decode("utf-8")
    download_url = f"data:{data_type};base64,{data}"

    return DownloadQACatalogResponse(
        download_url=download_url,
        filename=filename,
    )
